from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import current_user
from app.auth import role_required
from app.models import db, Class, Submission

faculty = Blueprint('faculty', __name__, url_prefix='/faculty')


@faculty.route('/dashboard')
@role_required('faculty')
def dashboard():
    """Faculty dashboard — overview of all classes."""
    classes = Class.query.filter_by(faculty_id=current_user.id).order_by(Class.created_at.desc()).all()

    total = len(classes)
    open_count = sum(1 for c in classes if c.is_open)
    closed_count = total - open_count

    return render_template('faculty/dashboard.html',
                           classes=classes,
                           total=total,
                           open_count=open_count,
                           closed_count=closed_count)


@faculty.route('/class/<string:code>')
@role_required('faculty')
def class_detail(code):
    """Class detail page — view submissions, run comparison."""
    cls = Class.query.filter_by(code=code, faculty_id=current_user.id).first_or_404()
    submissions = Submission.query.filter_by(class_id=cls.id).order_by(Submission.uploaded_at.asc()).all()
    student_count = len(cls.students)

    return render_template('faculty/class_detail.html',
                           cls=cls,
                           submissions=submissions,
                           student_count=student_count)


@faculty.route('/class/create', methods=['GET', 'POST'])
@role_required('faculty')
def create_class():
    """Create a new class — placeholder, built in Class Management module."""
    from flask import flash
    import string, random
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Class name is required.', 'error')
            return render_template('faculty/create_class.html')

        # Ensure the generated class code is unique to avoid IntegrityError
        while True:
            code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            if not Class.query.filter_by(code=code).first():
                break

        new_class = Class(name=name, code=code, faculty_id=current_user.id)
        db.session.add(new_class)
        db.session.commit()
        flash(f'Class "{name}" created! Code: {code}', 'success')
        return redirect(url_for('faculty.dashboard'))
    return render_template('faculty/create_class.html')


@faculty.route('/class/<string:code>/close', methods=['POST'])
@role_required('faculty')
def close_uploads(code):
    """Close uploads for a class."""
    from flask import flash
    cls = Class.query.filter_by(code=code, faculty_id=current_user.id).first_or_404()
    cls.is_open = False
    db.session.commit()
    flash(f'Uploads closed for "{cls.name}".', 'success')
    return redirect(url_for('faculty.class_detail', code=code))


@faculty.route('/class/<string:code>/reopen', methods=['POST'])
@role_required('faculty')
def reopen_uploads(code):
    """Reopen uploads for a closed class."""
    from flask import flash
    cls = Class.query.filter_by(code=code, faculty_id=current_user.id).first_or_404()
    cls.is_open = True
    db.session.commit()
    flash(f'Uploads reopened for "{cls.name}".', 'success')
    return redirect(url_for('faculty.class_detail', code=code))


@faculty.route('/class/<string:code>/compare', methods=['POST'])
@role_required('faculty')
def run_comparison(code):
    """Run similarity comparison using TF-IDF + Cosine Similarity."""
    cls = Class.query.filter_by(code=code, faculty_id=current_user.id).first_or_404()

    if cls.is_open:
        flash('Close uploads before running comparison.', 'error')
        return redirect(url_for('faculty.class_detail', code=code))

    from app.engine import run_comparison as compare
    result = compare(cls.id)

    if 'error' in result:
        flash(result['error'], 'error')
    else:
        flash(
            f'Comparison complete! {result["total_pairs"]} pairs analysed, '
            f'{result["flagged_pairs"]} flagged (≥{int(result["threshold"]*100)}% similar).',
            'success'
        )
    return redirect(url_for('faculty.class_detail', code=code))


@faculty.route('/download/<int:submission_id>')
@role_required('faculty')
def download_file(submission_id):
    """Download a student's uploaded file."""
    import os
    from flask import send_file, abort, current_app
    sub = Submission.query.get_or_404(submission_id)
    cls = Class.query.filter_by(id=sub.class_id, faculty_id=current_user.id).first_or_404()
    if not sub.file_path or not os.path.exists(sub.file_path):
        flash('File not found.', 'error')
        return redirect(url_for('faculty.class_detail', code=cls.code))

    # Validate the file is within the expected upload directory
    real_path = os.path.realpath(sub.file_path)
    upload_root = os.path.realpath(current_app.config['UPLOAD_FOLDER'])
    if not real_path.startswith(upload_root + os.sep):
        abort(403)

    return send_file(real_path, as_attachment=True)


@faculty.route('/class/<string:code>/export', methods=['POST'])
@role_required('faculty')
def export_excel(code):
    """Export comparison results as an Excel file."""
    import io
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.models import Result

    cls = Class.query.filter_by(code=code, faculty_id=current_user.id).first_or_404()
    submissions = Submission.query.filter_by(class_id=cls.id).order_by(Submission.uploaded_at.asc()).all()

    # Build collision map: for each rejected sub, find who they collided with
    collision_map = {}
    for sub in submissions:
        if sub.status == 'rejected':
            # Find the highest-score result involving this submission
            result = Result.query.filter(
                Result.class_id == cls.id,
                Result.is_similar == True,
                ((Result.submission_2_id == sub.id))
            ).order_by(Result.similarity_score.desc()).first()

            if result:
                # The other submission (the one that was accepted first)
                other = Submission.query.get(result.submission_1_id)
                if other:
                    collision_map[sub.id] = other.student.name

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Comparison Report'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    accepted_font = Font(color='166534', bold=True)
    accepted_fill = PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid')
    rejected_font = Font(color='991b1b', bold=True)
    rejected_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:C1')
    ws['A1'] = f'{cls.name} — Comparison Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Student Name', 'Project Title', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for i, sub in enumerate(submissions, 4):
        ws.cell(row=i, column=1, value=sub.student.name).border = thin_border
        ws.cell(row=i, column=2, value=sub.project_title).border = thin_border

        status_cell = ws.cell(row=i, column=3)
        status_cell.border = thin_border

        if sub.status == 'accepted':
            status_cell.value = 'Accepted'
            status_cell.font = accepted_font
            status_cell.fill = accepted_fill
        elif sub.status == 'rejected':
            collided_with = collision_map.get(sub.id, 'unknown')
            status_cell.value = f'Collision with {collided_with}'
            status_cell.font = rejected_font
            status_cell.fill = rejected_fill
        else:
            status_cell.value = sub.status.capitalize()

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 40

    # Save to memory and send
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'{cls.name.replace(" ", "_")}_Report.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@faculty.route('/class/<string:code>/delete', methods=['POST'])
@role_required('faculty')
def delete_class(code):
    """Delete a closed class and all its data, including files on disk."""
    import os, shutil
    from flask import current_app
    from app.models import Result, class_students
    cls = Class.query.filter_by(code=code, faculty_id=current_user.id).first_or_404()
    if cls.is_open:
        flash('Cannot delete an open class. Close it first.', 'error')
        return redirect(url_for('faculty.dashboard'))
    # Delete uploaded files from disk
    upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], cls.code)
    if os.path.exists(upload_dir):
        shutil.rmtree(upload_dir)
    # Delete in order: results → submissions → student associations → class
    Result.query.filter_by(class_id=cls.id).delete()
    Submission.query.filter_by(class_id=cls.id).delete()
    db.session.execute(class_students.delete().where(class_students.c.class_id == cls.id))
    db.session.delete(cls)
    db.session.commit()
    flash(f'Class "{cls.name}" deleted.', 'success')
    return redirect(url_for('faculty.dashboard'))

